from datetime import datetime
import os
from openpyxl import load_workbook
from db import get_db_connection
from config import FIELD_MAP
from sqlite3 import IntegrityError

def normalize_header(h):
    if not h:
        return ""
    h = str(h).strip().lower()
    for ch in " -–—":
        h = h.replace(ch, "_")
    return h

def process_xlsx_file(file_path):
    """
    Читает XLSX, вставляет данные в SQLite, затем удаляет файл.
    Возвращает (success: bool, message: str)
    """
    upload_time = datetime.utcnow().isoformat()

    try:
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        ws = wb.active

        if ws.max_row < 2:
            return False, "Файл слишком короткий: нет данных для загрузки"

        headers_raw = [cell.value for cell in ws[1]]
        headers = [normalize_header(h) for h in headers_raw]

        mapping = {field: None for field in FIELD_MAP}
        for idx, h in enumerate(headers):
            for field, keywords in FIELD_MAP.items():
                if mapping[field] is not None:
                    continue
                for kw in keywords:
                    if kw in h:
                        mapping[field] = idx
                        break

        missing = [f for f, idx in mapping.items() if idx is None]
        if missing:
            return False, f"Не найдены колонки для полей: {', '.join(missing)}. Проверьте заголовки в Excel."

        conn = get_db_connection()
        cur = conn.cursor()
        insert_sql = """
            INSERT INTO shipments (shpi, mass, shipping_cost, recipient, phone, index_code, address, internal_number, comment, uploaded_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        count = 0
        duplicates = 0
        errors = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                shpi_val = row[mapping["shpi"]].value
                if shpi_val is None or str(shpi_val).strip() == "":
                    continue
                shpi = str(shpi_val).strip()

                mass = row[mapping["mass"]].value
                shipping_cost = row[mapping["shipping_cost"]].value

                recipient = str(row[mapping["recipient"]].value).strip() if row[mapping["recipient"]].value else ""
                phone = str(row[mapping["phone"]].value).strip() if row[mapping["phone"]].value else ""
                index_code = str(row[mapping["index_code"]].value).strip() if row[mapping["index_code"]].value else ""
                address = str(row[mapping["address"]].value).strip() if row[mapping["address"]].value else ""
                internal_number = str(row[mapping["internal_number"]].value).strip() if row[mapping["internal_number"]].value else ""
                comment = str(row[mapping["comment"]].value).strip() if row[mapping["comment"]].value else ""

                cur.execute(insert_sql, (shpi, mass, shipping_cost, recipient, phone, index_code, address, internal_number, comment, upload_time))
                count += 1

            except IntegrityError:
                duplicates += 1
                continue
            except Exception as e:
                errors += 1
                print(f"Ошибка строки {row_idx} в файле {file_path}: {e}")

        conn.commit()
        conn.close()
        wb.close()

        os.remove(file_path)
        msg = f"Обработано строк: {count}, пропущено дублей: {duplicates}, ошибок: {errors}"
        print(f"[WATCHDOG] Успешно: {msg}")
        return True, msg

    except Exception as e:
        msg = str(e)
        print(f"[WATCHDOG] Ошибка обработки файла {file_path}: {msg}")
        # Файл НЕ удаляем при ошибке
        return False, msg
